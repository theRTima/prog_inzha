import os
import csv
from datetime import datetime
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTextEdit, 
                             QPushButton, QLabel, QComboBox, QDateEdit, 
                             QMessageBox, QTabWidget, QTableWidget, 
                             QTableWidgetItem, QHeaderView, QFileDialog)
from PyQt5.QtCore import QDate, Qt
from PyQt5.QtGui import QFont
from config.database import get_db
from repositories.order_repository import OrderRepository
from repositories.inventory_repository import InventoryRepository

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ReportsTab(QWidget):
    def __init__(self):
        super().__init__()
        self.current_report_data = None
        self.current_report_type = None
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout(self)

        # Панель управления отчетами
        control_panel = QHBoxLayout()

        self.report_type = QComboBox()
        self.report_type.addItems([
            'Отчет по продажам',
            'Отчет по остаткам',
            'Отчет по заказам'
        ])

        self.date_from = QDateEdit()
        self.date_from.setDate(QDate.currentDate().addDays(-30))
        self.date_from.setCalendarPopup(True)
        self.date_from.setDisplayFormat("dd.MM.yyyy")

        self.date_to = QDateEdit()
        self.date_to.setDate(QDate.currentDate())
        self.date_to.setCalendarPopup(True)
        self.date_to.setDisplayFormat("dd.MM.yyyy")

        btn_generate = QPushButton('Сгенерировать отчет')
        btn_export = QPushButton('Экспорт в файл')

        control_panel.addWidget(QLabel('Тип отчета:'))
        control_panel.addWidget(self.report_type)
        control_panel.addWidget(QLabel('С:'))
        control_panel.addWidget(self.date_from)
        control_panel.addWidget(QLabel('По:'))
        control_panel.addWidget(self.date_to)
        control_panel.addWidget(btn_generate)
        control_panel.addWidget(btn_export)
        control_panel.addStretch()

        layout.addLayout(control_panel)

        # Таблица для отображения отчета
        self.report_table = QTableWidget()
        self.report_table.setColumnCount(0)
        self.report_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.report_table)

        # Подключение сигналов
        btn_generate.clicked.connect(self.generate_report)
        btn_export.clicked.connect(self.export_report)

    def generate_report(self):
        """Генерация отчета"""
        report_type = self.report_type.currentText()
        self.current_report_type = report_type
        
        try:
            if report_type == 'Отчет по продажам':
                self.current_report_data = self.generate_sales_report()
            elif report_type == 'Отчет по остаткам':
                self.current_report_data = self.generate_inventory_report()
            elif report_type == 'Отчет по заказам':
                self.current_report_data = self.generate_orders_report()
                
        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', f'Не удалось сгенерировать отчет: {str(e)}')

    def generate_sales_report(self):
        """Генерация отчета по продажам"""
        start_date = datetime.combine(self.date_from.date().toPyDate(), datetime.min.time())
        end_date = datetime.combine(self.date_to.date().toPyDate(), datetime.max.time())
        
        with get_db() as db:
            repo = OrderRepository(db)
            report_data = repo.get_sales_report(start_date, end_date)
        
        # Настраиваем таблицу
        self.report_table.setRowCount(0)
        self.report_table.setColumnCount(4)
        self.report_table.setHorizontalHeaderLabels([
            'Показатель', 'Значение', 'Детали', 'Примечание'
        ])
        
        row = 0
        
        # Шапка отчета
        self.add_table_row(row, 'ОТЧЕТ ПО ПРОДАЖАМ', '', '', True)
        row += 1
        self.add_table_row(row, 'Период:', 
                          f"{report_data['period_start'].strftime('%d.%m.%Y')} - {report_data['period_end'].strftime('%d.%m.%Y')}", '', True)
        row += 1
        self.add_table_row(row, 'Дата формирования:', 
                          datetime.now().strftime('%d.%m.%Y %H:%M'), '', True)
        row += 1
        
        # Основные показатели
        self.add_table_row(row, 'Общая выручка:', 
                          f"{report_data['total_revenue']:,.2f} руб.", '', True)
        row += 1
        self.add_table_row(row, 'Количество заказов:', 
                          str(report_data['total_orders']), '', True)
        row += 1
        
        # Статистика по дням
        if report_data['daily_stats']:
            self.add_table_row(row, 'Статистика по дням:', '', '', True)
            row += 1
            
            for daily in report_data['daily_stats']:
                self.add_table_row(row, 
                                  daily.date.strftime('%d.%m.%Y'),
                                  f"{daily.daily_revenue:,.2f} руб.",
                                  f"{daily.orders_count} заказов",
                                  False)
                row += 1
        
        # Статистика по блюдам
        if report_data['dishes_stats']:
            self.add_table_row(row, 'Топ блюд по выручке:', '', '', True)
            row += 1
            
            for dish in report_data['dishes_stats']:
                self.add_table_row(row,
                                  f"{dish.name} ({dish.category})",
                                  f"{dish.total_revenue:,.2f} руб.",
                                  f"{dish.total_quantity} шт.",
                                  False)
                row += 1
        
        self.report_table.resizeRowsToContents()
        return report_data

    def generate_inventory_report(self):
        """Генерация отчета по остаткам"""
        with get_db() as db:
            repo = InventoryRepository(db)
            report_data = repo.get_inventory_report()
        
        # Настраиваем таблицу
        self.report_table.setRowCount(0)
        self.report_table.setColumnCount(5)
        self.report_table.setHorizontalHeaderLabels([
            'Категория', 'Наименование', 'Текущий остаток', 'Мин. остаток', 'Статус'
        ])
        
        row = 0
        
        # Шапка отчета
        self.add_table_row(row, 'ОТЧЕТ ПО ОСТАТКАМ НА СКЛАДЕ', '', '', '', True)
        row += 1
        self.add_table_row(row, 'Дата формирования:', 
                          report_data['report_date'].strftime('%d.%m.%Y %H:%M'), '', '', True)
        row += 1
        self.add_table_row(row, 'Всего позиций:', 
                          str(report_data['total_items']), '', '', True)
        row += 1
        self.add_table_row(row, 'Низкие остатки:', 
                          str(report_data['low_stock_count']), '', '', True)
        row += 1
        
        # Список позиций с низким остатком
        if report_data['low_stock_items']:
            self.add_table_row(row, 'ПОЗИЦИИ С НИЗКИМИ ОСТАТКАМИ:', '', '', '', True)
            row += 1
            
            for item in report_data['low_stock_items']:
                status = 'НИЗКИЙ ОСТАТОК'
                self.add_table_row(row,
                                  item['category'],
                                  item['name'],
                                  f"{item['current_stock']:.3f} {item['unit']}",
                                  f"{item['min_stock']:.3f} {item['unit']}",
                                  status,
                                  True)
                row += 1
        
        # Все позиции по категориям
        self.add_table_row(row, 'ВСЕ ПОЗИЦИИ НА СКЛАДЕ:', '', '', '', True)
        row += 1
        
        for category, stats in report_data['category_stats'].items():
            self.add_table_row(row, f"Категория: {category} ({stats['count']} позиций)", '', '', '', True)
            row += 1
            
            for item in stats['items']:
                status = 'НОРМА' if item['current_stock'] >= item['min_stock'] else 'НИЗКИЙ'
                self.add_table_row(row,
                                  '',
                                  item['name'],
                                  f"{item['current_stock']:.3f} {item['unit']}",
                                  f"{item['min_stock']:.3f} {item['unit']}",
                                  status,
                                  False)
                row += 1
        
        self.report_table.resizeRowsToContents()
        return report_data

    def generate_orders_report(self):
        """Генерация отчета по заказам"""
        start_date = datetime.combine(self.date_from.date().toPyDate(), datetime.min.time())
        end_date = datetime.combine(self.date_to.date().toPyDate(), datetime.max.time())
        
        with get_db() as db:
            repo = OrderRepository(db)
            orders = repo.get_orders_report(start_date, end_date)
        
        # Настраиваем таблицу
        self.report_table.setRowCount(0)
        self.report_table.setColumnCount(7)
        self.report_table.setHorizontalHeaderLabels([
            'ID заказа', 'Клиент', 'Телефон', 'Дата/время', 'Статус', 'Сумма', 'Примечание'
        ])
        
        row = 0
        
        # Шапка отчета
        self.add_table_row(row, 'ОТЧЕТ ПО ЗАКАЗАМ', '', '', '', '', '', True)
        row += 1
        self.add_table_row(row, 'Период:', 
                          f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}", 
                          '', '', '', '', True)
        row += 1
        self.add_table_row(row, 'Дата формирования:', 
                          datetime.now().strftime('%d.%m.%Y %H:%M'), '', '', '', '', True)
        row += 1
        self.add_table_row(row, 'Всего заказов:', str(len(orders)), '', '', '', '', True)
        row += 1
        
        # Список заказов
        total_revenue = 0
        for order in orders:
            total_revenue += order['total']
            
            self.add_table_row(row,
                              str(order['id']),
                              order['customer_name'],
                              order['phone'],
                              order['created'].strftime('%d.%m.%Y %H:%M'),
                              order['status'],
                              f"{order['total']:,.2f} руб.",
                              order['notes'] or '',
                              True)
            row += 1
            
            # Детали заказа
            if order['items']:
                for item in order['items']:
                    self.add_table_row(row,
                                      '',
                                      f"  - {item['dish_name']}",
                                      f"x{item['quantity']}",
                                      '',
                                      '',
                                      f"{item['subtotal']:,.2f} руб.",
                                      '',
                                      False)
                    row += 1
        
        # Итоги
        self.add_table_row(row, 'ИТОГО:', '', '', '', '', f"{total_revenue:,.2f} руб.", '', True)
        
        self.report_table.resizeRowsToContents()
        return orders

    def add_table_row(self, row, *values, header=False):
        """Добавляет строку в таблицу отчета"""
        self.report_table.insertRow(row)
        
        for col, value in enumerate(values):
            item = QTableWidgetItem(str(value))
            
            if header:
                item.setFont(QFont('Arial', 10, QFont.Bold))
                item.setBackground(Qt.lightGray)
            
            self.report_table.setItem(row, col, item)

    def export_report(self):
        """Экспорт отчета в файл"""
        if self.report_table.rowCount() == 0:
            QMessageBox.warning(self, 'Предупреждение', 'Сначала сгенерируйте отчет')
            return
        
        # Определяем доступные форматы
        formats = []
        filters = []
        
        # Всегда доступен CSV
        formats.append('CSV (*.csv)')
        filters.append('csv')
        
        # Проверяем доступность других форматов
        if PANDAS_AVAILABLE or OPENPYXL_AVAILABLE:
            formats.append('Excel (*.xlsx)')
            filters.append('xlsx')
        
        filter_str = ';;'.join(formats)
        
        # Открываем диалог сохранения файла
        file_path, selected_filter = QFileDialog.getSaveFileName(
            self,
            'Экспорт отчета',
            f"{self.current_report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            filter_str
        )
        
        if not file_path:
            return
        
        # Определяем формат по расширению
        if file_path.lower().endswith('.csv'):
            self.export_to_csv(file_path)
        elif file_path.lower().endswith('.xlsx'):
            self.export_to_excel(file_path)
        else:
            # Если расширение не указано, добавляем по умолчанию
            if selected_filter:
                if 'csv' in selected_filter.lower():
                    file_path += '.csv'
                    self.export_to_csv(file_path)
                elif 'xlsx' in selected_filter.lower():
                    file_path += '.xlsx'
                    self.export_to_excel(file_path)

    def export_to_csv(self, file_path):
        """Экспорт в CSV"""
        try:
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                # Записываем заголовки
                headers = []
                for col in range(self.report_table.columnCount()):
                    header_item = self.report_table.horizontalHeaderItem(col)
                    if header_item:
                        headers.append(header_item.text())
                    else:
                        headers.append(f'Столбец {col+1}')
                writer.writerow(headers)
                
                # Записываем данные
                for row in range(self.report_table.rowCount()):
                    row_data = []
                    for col in range(self.report_table.columnCount()):
                        item = self.report_table.item(row, col)
                        if item:
                            # Заменяем разделители в тексте, чтобы не портить CSV
                            text = item.text().replace(';', ',').replace('\n', ' ')
                            row_data.append(text)
                        else:
                            row_data.append('')
                    writer.writerow(row_data)
            
            QMessageBox.information(self, 'Успех', f'Отчет сохранен в файл:\n{file_path}')
            
        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', f'Не удалось сохранить отчет в CSV: {str(e)}')

    def export_to_excel(self, file_path):
        """Экспорт в Excel"""
        try:
            if PANDAS_AVAILABLE:
                # Используем pandas если доступен
                self.export_to_excel_pandas(file_path)
            elif OPENPYXL_AVAILABLE:
                # Используем openpyxl если доступен
                self.export_to_excel_openpyxl(file_path)
            else:
                QMessageBox.warning(self, 'Ошибка', 
                                  'Для экспорта в Excel требуется установить библиотеку pandas или openpyxl.\n\n'
                                  'Установите одну из команд:\n'
                                  'pip install pandas\n'
                                  'pip install openpyxl')
                return
                
        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', f'Не удалось сохранить отчет в Excel: {str(e)}')

    def export_to_excel_pandas(self, file_path):
        """Экспорт в Excel через pandas"""
        # Собираем данные из таблицы
        data = []
        
        # Заголовки
        headers = []
        for col in range(self.report_table.columnCount()):
            header_item = self.report_table.horizontalHeaderItem(col)
            if header_item:
                headers.append(header_item.text())
            else:
                headers.append(f'Столбец {col+1}')
        
        # Данные
        for row in range(self.report_table.rowCount()):
            row_data = []
            for col in range(self.report_table.columnCount()):
                item = self.report_table.item(row, col)
                if item:
                    row_data.append(item.text())
                else:
                    row_data.append('')
            data.append(row_data)
        
        # Создаем DataFrame
        df = pd.DataFrame(data, columns=headers)
        
        # Сохраняем в Excel
        df.to_excel(file_path, index=False, engine='openpyxl')
        QMessageBox.information(self, 'Успех', f'Отчет сохранен в файл:\n{file_path}')

    def export_to_excel_openpyxl(self, file_path):
        """Экспорт в Excel через openpyxl"""
        wb = Workbook()
        ws = wb.active
        ws.title = self.current_report_type if self.current_report_type else 'Отчет'
        
        # Записываем заголовки
        for col in range(self.report_table.columnCount()):
            header_item = self.report_table.horizontalHeaderItem(col)
            if header_item:
                ws.cell(row=1, column=col+1, value=header_item.text())
        
        # Записываем данные
        for row in range(self.report_table.rowCount()):
            for col in range(self.report_table.columnCount()):
                item = self.report_table.item(row, col)
                if item:
                    ws.cell(row=row+2, column=col+1, value=item.text())
        
        # Автоматическая ширина столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
        QMessageBox.information(self, 'Успех', f'Отчет сохранен в файл:\n{file_path}')